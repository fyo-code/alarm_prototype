"""ALARM V1 — API + static host.

Run from the alarm_v1 folder:
    .venv/bin/uvicorn app.server:app --port 8700 --reload
"""
from __future__ import annotations

import io
import json
import math
import re
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from fastapi import FastAPI, HTTPException, Query
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from pipeline import build as build_mod  # noqa: E402
from pipeline import paths  # noqa: E402

WEB = Path(__file__).resolve().parent / "web"

app = FastAPI(title="Alarm V1", docs_url="/api/docs")

COLOUR_ORDER = ["green", "orange", "red", "blue", "grey", "black"]

_state: dict = {}


# ------------------------------------------------------------------- state --

def load_state() -> None:
    if not paths.PANEL_PARQUET.exists():
        build_mod.build()
    panel = pd.read_parquet(paths.PANEL_PARQUET)
    meta = json.loads(paths.META_JSON.read_text(encoding="utf-8"))
    store = (pd.read_parquet(paths.OUT_DIR / "store_stock.parquet")
             if (paths.OUT_DIR / "store_stock.parquet").exists() else pd.DataFrame())
    _state["panel"] = panel
    _state["meta"] = meta
    _state["store"] = store
    _state["cfg"] = paths.load_config()


def P() -> pd.DataFrame:
    if "panel" not in _state:
        load_state()
    return _state["panel"]


def M() -> dict:
    if "meta" not in _state:
        load_state()
    return _state["meta"]


def _f(x) -> float:
    try:
        v = float(x)
    except (TypeError, ValueError):
        return 0.0
    return 0.0 if not math.isfinite(v) else v


def _clean(obj):
    """JSON-safe: numpy scalars out, NaN/Inf to None."""
    if isinstance(obj, dict):
        return {k: _clean(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple, np.ndarray)):
        return [_clean(v) for v in obj]
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating, float)):
        v = float(obj)
        return None if not math.isfinite(v) else round(v, 4)
    if isinstance(obj, (np.bool_,)):
        return bool(obj)
    if isinstance(obj, pd.Timestamp):
        return str(obj.date())
    if obj is pd.NaT or obj is pd.NA:
        return None
    if isinstance(obj, float) and math.isnan(obj):
        return None
    # any remaining pandas scalar NA (nullable string/boolean/Int) — these reach
    # here from warehouse-only rows and used to raise
    # "Object of type NAType is not JSON serializable" on /api/sku/{sku}
    try:
        if not isinstance(obj, (str, bytes, list, dict, tuple)) and pd.isna(obj):
            return None
    except (TypeError, ValueError):
        pass
    return obj


# -------------------------------------------------------------- aggregates --

def segment_stats(df: pd.DataFrame) -> list[dict]:
    tot_u = _f(df["stock_units"].sum())
    tot_v = _f(df["stock_value"].sum())
    rows = []
    for c in COLOUR_ORDER:
        sub = df[df["colour"] == c]
        if not len(sub):
            continue
        subs = (sub.groupby("substate")
                .agg(units=("stock_units", "sum"), value=("stock_value", "sum"),
                     skus=("sku", "count")).reset_index())
        rows.append({
            "colour": c,
            "units": _f(sub["stock_units"].sum()),
            "value": _f(sub["stock_value"].sum()),
            "skus": int(len(sub)),
            "unit_share": _f(sub["stock_units"].sum() / tot_u * 100) if tot_u else 0.0,
            "value_share": _f(sub["stock_value"].sum() / tot_v * 100) if tot_v else 0.0,
            "substates": [{"substate": r["substate"], "units": _f(r["units"]),
                           "value": _f(r["value"]), "skus": int(r["skus"])}
                          for _, r in subs.iterrows()],
        })
    return rows


def factory_table(df: pd.DataFrame) -> pd.DataFrame:
    stocked = df[df["stock_units"] > 0]
    g = stocked.groupby("factory").agg(
        units=("stock_units", "sum"),
        value=("stock_value", "sum"),
        skus=("sku", "count"),
    )
    dead = (stocked[stocked["colour"] == "grey"].groupby("factory")
            .agg(dead_units=("stock_units", "sum"), dead_value=("stock_value", "sum")))
    prop = (df[df["suggested_qty"] > 0].groupby("factory")
            .agg(reorder_qty=("suggested_qty", "sum"),
                 reorder_value=("suggested_value", "sum"),
                 reorder_skus=("sku", "count"),
                 priority=("priority_score", "sum")))
    moving = (stocked[stocked["colour"].isin(["green", "orange", "red", "blue"])]
              .groupby("factory")["stock_units"].sum().rename("moving_units"))
    # outer on `prop`: a factory can have a proposal while holding zero stock
    # (everything sold out). Dropping it would silently shrink the Level-0 total.
    out = (g.join(dead, how="left").join(prop, how="outer").join(moving, how="left")
           .fillna(0.0))
    out["dead_pct"] = np.where(out["units"] > 0, out["dead_units"] / out["units"] * 100, 0.0)
    out["moving_pct"] = np.where(out["units"] > 0, out["moving_units"] / out["units"] * 100, 0.0)
    return out.reset_index()


def rows_out(df: pd.DataFrame, cols: list[str]) -> list[dict]:
    recs = df[cols].to_dict("records")
    for r in recs:
        if isinstance(r.get("why"), str):
            try:
                r["why"] = json.loads(r["why"])
            except json.JSONDecodeError:
                r["why"] = []
    return recs


def coverage_note(part: float, whole: float) -> dict:
    return {"part": _f(part), "whole": _f(whole),
            "pct": _f(part / whole * 100) if whole else 0.0}


# -------------------------------------------------------------------- API ---

@app.get("/api/meta")
def api_meta():
    m = M()
    p = P()
    wh = m["coverage"]["warehouse"]
    unmapped_units = _f(wh.get("units_total", 0) - wh.get("units_matched", 0))
    unmapped_value = _f(wh.get("value_total", 0) - wh.get("value_matched", 0))
    return JSONResponse(_clean({
        "as_of": m["as_of"],
        "built_at": m["built_at"],
        "category": m["category"],
        "config": _state["cfg"],
        "totals": {
            "units": _f(p["stock_units"].sum()),
            "value": _f(p["stock_value"].sum()),
            "skus": int(len(p)),
            "skus_with_stock": int((p["stock_units"] > 0).sum()),
            "factories": int(p.loc[p["stock_units"] > 0, "factory"].nunique()),
            "families": int(p.loc[p["stock_units"] > 0, "family"].nunique()),
        },
        "coverage": m["coverage"],
        "unmapped": {"units": unmapped_units, "value": unmapped_value,
                     # the coded loader emits `excluded_non_rug`; the older
                     # name-keyed fallback emits `unmapped_buckets`
                     "buckets": (wh.get("excluded_non_rug")
                                 or wh.get("unmapped_buckets") or []),
                     "kind": ("excluded_non_rug" if wh.get("excluded_non_rug")
                              else "unmapped_by_name")},
        "cost_source_mix": m["cost_source_mix"],
        "state_source_mix": m["state_source_mix"],
        "lead_time_sources": m["lead_time_sources"],
        "factory_source_mix": m.get("factory_source_mix", {}),
        "product_state_mix": m.get("product_state_mix", {}),
    }))


@app.get("/api/level0")
def api_level0(family: str | None = None, top_n: int = 10, basis: str = "units"):
    """`basis` picks which measure the top-N is RANKED and MEASURED by.

    Previously the top-N was always chosen and ranked by units, and the coverage
    note always divided by the unit total, while the UI's "Lei" toggle only
    restyled the bar lengths. The result was a money screen showing the top ten
    factories by volume: 6 of 10 matched the real value top ten, covering 69.9%
    of the money while the note claimed 75.9% (the unit figure). A basis toggle
    has to change the selection, not just the drawing.
    """
    p = P()
    if family:
        p = p[p["family"] == family]
    by_value = basis == "value"
    ft = factory_table(p)
    tot_u = _f(p["stock_units"].sum())
    tot_v = _f(p["stock_value"].sum())

    stock_key = "value" if by_value else "units"
    reorder_key = "reorder_value" if by_value else "reorder_qty"
    by_reorder = ft.sort_values(reorder_key, ascending=False).head(top_n)
    by_units = ft.sort_values(stock_key, ascending=False).head(top_n)
    all_reorder_qty = _f(ft[reorder_key].sum())

    return JSONResponse(_clean({
        "segments": segment_stats(p),
        "totals": {"units": tot_u, "value": tot_v, "skus": int(len(p)),
                   "skus_with_stock": int((p["stock_units"] > 0).sum()),
                   "factories": int(ft.shape[0])},
        "top_reorder": {
            "rows": by_reorder.to_dict("records"),
            "coverage": coverage_note(by_reorder[reorder_key].sum(), all_reorder_qty),
            "n_total": int((ft["reorder_qty"] > 0).sum()),
        },
        "top_stock": {
            "rows": by_units.to_dict("records"),
            "coverage": coverage_note(by_units[stock_key].sum(), ft[stock_key].sum()),
            "n_total": int(ft.shape[0]),
        },
        "reorder_total": {
            "qty": all_reorder_qty,
            "value": _f(ft["reorder_value"].sum()),
            "skus": int((p["suggested_qty"] > 0).sum()),
        },
        "families_total": int(p.loc[p["stock_units"] > 0, "family"].nunique()),
        "basis": basis,
    }))


@app.get("/api/factory/{name}")
def api_factory(name: str, top_n: int = 15, basis: str = "units"):
    p = P()
    f = p[p["factory"] == name]
    if not len(f):
        raise HTTPException(404, f"unknown factory: {name}")
    stocked = f[f["stock_units"] > 0]

    reorder = (f[f["suggested_qty"] > 0]
               .sort_values("priority_score", ascending=False)
               .head(top_n))
    reorder_cols = ["sku", "denumire_articol", "family", "width_cm", "stock_units",
                    "months_of_supply", "days_of_cover", "rate_mo", "suggested_qty",
                    "suggested_value", "criticality", "priority_score", "colour",
                    "lead_time_days", "lead_time_source", "unit_cost", "why"]

    # per-colour breakdown in BOTH measures. Only units were returned before, so
    # in value mode the health composition collapsed into one flat blue segment
    # and the "where is the problem inside this family" question went unanswered.
    by_value = basis == "value"
    measure = "stock_value" if by_value else "stock_units"
    fam = (stocked.groupby(["family", "colour"])[measure].sum().unstack(fill_value=0.0))
    for c in COLOUR_ORDER:
        if c not in fam.columns:
            fam[c] = 0.0
    fam["total"] = fam[COLOUR_ORDER].sum(axis=1)
    fam = (fam.join(stocked.groupby("family")["stock_value"].sum().rename("value"))
              .join(stocked.groupby("family")["stock_units"].sum().rename("units"))
              .sort_values("total", ascending=False))
    fam_top = fam.head(top_n).reset_index()

    dead = (stocked[stocked["colour"] == "grey"]
            .groupby(["family", "substate"])
            .agg(units=("stock_units", "sum"), value=("stock_value", "sum"),
                 skus=("sku", "count")).reset_index())
    dead_piv = (dead.pivot_table(index="family", columns="substate",
                                 values=["units", "value"], aggfunc="sum", fill_value=0.0))
    dead_rows = []
    if len(dead_piv):
        for famname, row in dead_piv.iterrows():
            active_u = _f(row.get(("units", "grey_active"), 0))
            phase_u = _f(row.get(("units", "grey_phaseout"), 0))
            dead_rows.append({
                "family": famname,
                "dead_active_units": active_u,
                "dead_phaseout_units": phase_u,
                "units": active_u + phase_u,
                "value": _f(row.get(("value", "grey_active"), 0)) + _f(row.get(("value", "grey_phaseout"), 0)),
            })
        dead_rows.sort(key=lambda r: -r["units"])
    dead_rows = dead_rows[:top_n]

    # worst performers — "măcar să știu ce să nu mai fac"
    worst = (stocked.groupby("family")
             .agg(units=("stock_units", "sum"), value=("stock_value", "sum"),
                  sold90=("units_90d", "sum"), sold365=("units_365d", "sum"),
                  skus=("sku", "count"),
                  dead_units=("stock_units", lambda s: 0.0)).reset_index())
    dead_by_fam = (stocked[stocked["colour"] == "grey"].groupby("family")["stock_units"]
                   .sum().rename("dead_u"))
    worst = worst.merge(dead_by_fam, on="family", how="left").fillna({"dead_u": 0.0})
    worst["dead_pct"] = np.where(worst["units"] > 0, worst["dead_u"] / worst["units"] * 100, 0.0)
    worst["rotation_months"] = np.where(worst["sold90"] > 0,
                                        worst["units"] / (worst["sold90"] / 3.0), np.inf)
    worst = worst[worst["units"] >= 5].copy()
    # trapped money is the ACTUAL value of the dead rows, not the family value
    # scaled by the dead UNIT share. The old estimate assumed dead and moving
    # units cost the same, which understated expensive dead stock badly
    # (GIZA/BEACON showed 9,860 lei against a real 32,397).
    dead_val = (stocked[stocked["colour"] == "grey"].groupby("family")["stock_value"]
                .sum().rename("trapped"))
    worst = worst.merge(dead_val, on="family", how="left")
    worst["trapped"] = worst["trapped"].fillna(0.0)
    worst = worst.sort_values(["trapped", "units"], ascending=False).head(top_n)
    worst = worst.drop(columns=["dead_units"])

    widths = (stocked.dropna(subset=["width_cm"])
              .groupby("width_cm")
              .agg(units=("stock_units", "sum"), value=("stock_value", "sum"),
                   skus=("sku", "count")).reset_index()
              .sort_values("units", ascending=False))

    tot_u = _f(stocked["stock_units"].sum())
    moving = _f(stocked.loc[stocked["colour"] != "grey", "stock_units"].sum())
    all_ft = factory_table(p)
    rank = all_ft.sort_values("units", ascending=False).reset_index(drop=True)
    my_rank = int(rank.index[rank["factory"] == name][0]) + 1 if (rank["factory"] == name).any() else None

    return JSONResponse(_clean({
        "factory": name,
        "headline": {
            "units": tot_u,
            "value": _f(stocked["stock_value"].sum()),
            "skus": int(len(stocked)),
            "families": int(stocked["family"].nunique()),
            "reorder_qty": _f(f["suggested_qty"].sum()),
            "reorder_value": _f(f["suggested_value"].sum()),
            "reorder_skus": int((f["suggested_qty"] > 0).sum()),
            "moving_pct": _f(moving / tot_u * 100) if tot_u else 0.0,
            "dead_pct": _f(100 - (moving / tot_u * 100)) if tot_u else 0.0,
            "lead_time_days": int(f["lead_time_days"].iloc[0]),
            "lead_time_source": f["lead_time_source"].iloc[0],
            "share_of_all_units": _f(tot_u / _f(p["stock_units"].sum()) * 100),
            "share_of_all_value": _f(_f(stocked["stock_value"].sum()) / _f(p["stock_value"].sum()) * 100),
            "rank_by_units": my_rank,
            "n_factories": int(len(rank)),
        },
        "segments": segment_stats(f),
        "reorder": rows_out(reorder, reorder_cols),
        "reorder_n_total": int((f["suggested_qty"] > 0).sum()),
        "families": fam_top.to_dict("records"),
        "families_n_total": int(stocked["family"].nunique()),
        "families_measure_total": _f(stocked[measure].sum()),
        "basis": basis,
        "dead_by_family": dead_rows,
        "worst": worst.to_dict("records"),
        "widths": widths.to_dict("records"),
    }))


@app.get("/api/skus")
def api_skus(factory: str | None = None, family: str | None = None,
             colour: str | None = None, substate: str | None = None,
             reorder_only: bool = False, q: str | None = None,
             sort: str = "priority_score", limit: int = 400):
    p = P()
    if factory:
        p = p[p["factory"] == factory]
    if family:
        p = p[p["family"] == family]
    if colour:
        p = p[p["colour"] == colour]
    if substate:
        p = p[p["substate"] == substate]
    if reorder_only:
        p = p[p["suggested_qty"] > 0]
    if q:
        ql = q.lower()
        # regex=False: the box is a literal search. Without it a user typing "["
        # or "(" raised ArrowInvalid and the whole page became an error screen,
        # and "." silently matched everything.
        p = p[p["sku"].str.lower().str.contains(ql, na=False, regex=False)
              | p["denumire_articol"].fillna("").str.lower().str.contains(ql, na=False, regex=False)
              | p["family"].fillna("").str.lower().str.contains(ql, na=False, regex=False)]
    if sort in p.columns:
        p = p.sort_values(sort, ascending=False, na_position="last")
    cols = ["sku", "denumire_articol", "factory", "family", "width_cm", "stock_units",
            "wh_units", "store_units", "unit_cost", "stock_value", "rate_mo",
            "months_of_supply", "days_of_cover", "months_since_sale", "units_90d",
            "units_365d", "suggested_qty", "suggested_value", "criticality",
            "priority_score", "colour", "substate", "product_state", "state_source",
            "lead_time_days", "lead_time_source", "safety_stock", "reorder_point",
            "target_stock", "sigma_mo", "simple_rotation_months", "cost_source",
            "producer_entity", "factory_source", "why"]
    return JSONResponse(_clean({
        "n_total": int(len(p)),
        "rows": rows_out(p.head(limit), cols),
    }))


@app.get("/api/sku/{sku}")
def api_sku(sku: str):
    p = P()
    row = p[p["sku"] == sku]
    if not len(row):
        raise HTTPException(404, sku)
    r = row.iloc[0]
    why = r["why"]
    if isinstance(why, str):
        try:
            why = json.loads(why)
        except json.JSONDecodeError:
            why = []
    store = _state["store"]
    per_store = (store[store["sku"] == sku][["store_code", "units"]].to_dict("records")
                 if len(store) else [])
    siblings = p[(p["family"] == r["family"]) & (p["sku"] != sku)]
    sib = siblings[["sku", "width_cm", "stock_units", "months_of_supply",
                    "suggested_qty", "colour"]].sort_values("width_cm").to_dict("records")
    return JSONResponse(_clean({
        "row": {**{c: r[c] for c in p.columns if c not in ("in_photo", "why")}, "why": why},
        "per_store": per_store,
        "family_siblings": sib,
    }))


@app.get("/api/families")
def api_families(factory: str | None = None, limit: int = 60):
    p = P()
    if factory:
        p = p[p["factory"] == factory]
    stocked = p[p["stock_units"] > 0]
    g = (stocked.groupby("family")
         .agg(units=("stock_units", "sum"), value=("stock_value", "sum"),
              skus=("sku", "count"), sold90=("units_90d", "sum"))
         .reset_index().sort_values("units", ascending=False))
    return JSONResponse(_clean({"rows": g.head(limit).to_dict("records"),
                                "n_total": int(len(g))}))


@app.get("/api/leadtimes")
def api_leadtimes():
    p = P()
    ft = factory_table(p).sort_values("units", ascending=False)
    lt = (p.groupby("factory")
          .agg(lead_time_days=("lead_time_days", "first"),
               lead_time_source=("lead_time_source", "first")).reset_index())
    out = ft.merge(lt, on="factory", how="left")
    return JSONResponse(_clean({"rows": out.to_dict("records"),
                                "default_days": _state["cfg"]["replenishment"]["default_lead_time_days"]}))


@app.post("/api/leadtimes")
def api_set_leadtimes(payload: dict):
    """Write factory lead times and rebuild. ~30 numbers is an acceptable ask;
    per-SKU input is not, and the pipeline never asks for it."""
    rows = payload.get("rows") or []
    if not rows:
        raise HTTPException(400, "no rows")
    df = pd.DataFrame(rows)
    if not {"factory", "lead_time_days"} <= set(df.columns):
        raise HTTPException(400, "need factory + lead_time_days")
    df["lead_time_source"] = df.get("lead_time_source", "ENTERED_IN_APP")
    df.loc[df["lead_time_source"].isna(), "lead_time_source"] = "ENTERED_IN_APP"
    df[["factory", "lead_time_days", "lead_time_source"]].to_csv(
        paths.MANUAL_DIR / "lead_times.csv", index=False)
    build_mod.build()
    load_state()
    return JSONResponse({"ok": True, "rebuilt": True})


@app.get("/api/sanity")
def api_sanity():
    """Two internal cross-checks. Both computed from our data only.

    1. `simple_months` — the plain rotation heuristic (total stock ÷ sales over
       three months) against `engine_months` (the blended, variability-aware
       rate the reorder logic actually uses). Where they diverge, the plain
       number is being fooled by seasonality or by one atypical quarter, and
       that is worth knowing before defending a quantity.

    2. `unit_vs_money` — unit share against value share per state. These are
       genuinely different pictures and the money one is what a budget
       conversation runs on.

    No external baseline is used here. Nothing in this endpoint compares our
    output to numbers produced elsewhere, on another period or another scope —
    such a comparison would not be like-for-like and would mislead.
    """
    p = P()
    stocked = p[p["stock_units"] > 0]
    g = (stocked.groupby("factory")
         .agg(units=("stock_units", "sum"), sold90=("units_90d", "sum"),
              value=("stock_value", "sum"), skus=("sku", "count")).reset_index())
    g["simple_months"] = np.where(g["sold90"] > 0, g["units"] / (g["sold90"] / 3.0), np.inf)
    eng = (stocked.groupby("factory")
           .apply(lambda d: _f((d["stock_units"].sum()) /
                               (d["rate_mo"].sum() if d["rate_mo"].sum() > 0 else np.nan)),
                  include_groups=False)
           .rename("engine_months").reset_index())
    g = g.merge(eng, on="factory", how="left")
    g["delta_pct"] = np.where((g["simple_months"] > 0) & np.isfinite(g["simple_months"]),
                              (g["engine_months"] - g["simple_months"]) / g["simple_months"] * 100,
                              np.nan)
    g = g.sort_values("units", ascending=False)

    tot_u = _f(stocked["stock_units"].sum())
    tot_v = _f(stocked["stock_value"].sum())
    tot_s90 = _f(stocked["units_90d"].sum())
    tot_rate = _f(stocked["rate_mo"].sum())

    segs = segment_stats(p)
    unit_vs_money = [{
        "colour": s["colour"],
        "unit_share": s["unit_share"],
        "value_share": s["value_share"],
        "gap_pp": _f(s["value_share"] - s["unit_share"]),
        "units": s["units"], "value": s["value"], "skus": s["skus"],
    } for s in segs if s["units"] > 0 or s["value"] > 0]

    return JSONResponse(_clean({
        "overall": {
            "simple_months": _f(tot_u / (tot_s90 / 3.0)) if tot_s90 else None,
            "engine_months": _f(tot_u / tot_rate) if tot_rate else None,
            "units": tot_u, "value": tot_v, "sold90": tot_s90,
        },
        "by_factory": g.head(25).to_dict("records"),
        "unit_vs_money": unit_vs_money,
        "ours": {
            "units": tot_u,
            "value": tot_v,
            "skus": int((p["stock_units"] > 0).sum()),
            "segments": segs,
        },
    }))


# ----------------------------------------------------------------- exports --

def _excel(df: pd.DataFrame, sheet: str, title: str, qty_col: str | None,
           notes: list[str]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = sheet[:31]

    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = "MULTIPLICATOR / MULTIPLIER"
    ws["A2"].font = Font(bold=True)
    ws["B2"] = 1.0
    ws["B2"].fill = PatternFill("solid", fgColor="FFF3C4")
    ws["B2"].font = Font(bold=True)
    ws["C2"] = ("Schimbă doar celula B2 — toate cantităţile comandate se recalculează. "
                "/ Change only cell B2 — every order quantity rescales.")
    ws["C2"].alignment = Alignment(horizontal="left")

    header_row = 4
    cols = list(df.columns)
    if qty_col and qty_col in cols:
        cols = cols + [f"{qty_col}_x_multiplicator"]
    for j, c in enumerate(cols, start=1):
        cell = ws.cell(row=header_row, column=j, value=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F3A46")
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    for i, (_, r) in enumerate(df.iterrows(), start=header_row + 1):
        for j, c in enumerate(df.columns, start=1):
            v = r[c]
            # pd.NA reaches here from the nullable-string columns and openpyxl
            # rejects it outright ("Cannot convert <NA> to Excel")
            if v is None or (not isinstance(v, (list, dict, np.ndarray)) and pd.isna(v)):
                v = None
            elif isinstance(v, (np.integer,)):
                v = int(v)
            elif isinstance(v, (np.floating, float)):
                v = None if not math.isfinite(float(v)) else round(float(v), 2)
            elif isinstance(v, (list, dict)):
                v = json.dumps(v, ensure_ascii=False)[:2000]
            elif isinstance(v, pd.Timestamp):
                v = str(v.date())
            ws.cell(row=i, column=j, value=v)
        if qty_col and qty_col in df.columns:
            qcol = get_column_letter(list(df.columns).index(qty_col) + 1)
            ws.cell(row=i, column=len(cols),
                    value=f"=ROUND({qcol}{i}*$B$2,0)")

    for j, c in enumerate(cols, start=1):
        width = max(12, min(46, int(df[c].astype(str).str.len().max()) + 2)) if c in df.columns else 22
        ws.column_dimensions[get_column_letter(j)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    ns = wb.create_sheet("CITESTE_ME_README")
    ns["A1"] = "Ipoteze şi limitări / Assumptions and limitations"
    ns["A1"].font = Font(bold=True, size=13)
    for i, n in enumerate(notes, start=3):
        ns.cell(row=i, column=1, value=n).alignment = Alignment(wrap_text=True)
    ns.column_dimensions["A"].width = 120

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _export_notes() -> list[str]:
    m = M()
    cfg = _state["cfg"]
    wh = m["coverage"]["warehouse"]
    return [
        f"Fotografie la data / Photograph as of: {m['as_of']}. {cfg['as_of_note_ro']}",
        f"Stoc = depozit ({wh.get('period','?')}) + magazine. "
        f"{wh.get('unit_match_pct',0)}% din unităţile de depozit sunt mapate la un cod SKU; restul apare separat ca 'nemapat'.",
        f"Timp de livrare / lead time: {cfg['replenishment']['default_lead_time_days']} zile — "
        f"{cfg['replenishment']['default_lead_time_source']}. Timpul de livrare schimbă PRIORITATEA, nu cantitatea.",
        f"Stoc de siguranţă = z({cfg['replenishment']['service_level_pct']}%) × sigma lunar × "
        f"√(lead time + {cfg['replenishment']['review_period_months']} luni ciclu comandă). Formulă, niciodată introdusă manual.",
        f"Valoare = {cfg['money']['basis']}. {cfg['money']['basis_note_ro']}",
        "Starea produsului (in/out/phase-out) NU există în export. Unde apare, este un PROXY EURISTIC derivat "
        "din comportamentul stocului — de înlocuit cu exportul real.",
        "Capcană 1 — vânzări pe cod diferit: marfa vândută sub alt cod apare aici ca zero vânzări. "
        "Un produs care se vinde bine poate arăta mort.",
        "Capcană 2 — consum intern / proiecte: marfa consumată pentru amenajări sau proiecte apare ca cerere de retail. "
        "Poate produce comenzi false.",
        "Cantităţile sunt o BAZĂ. Proporţiile între SKU-uri sunt produsul; scala absolută rămâne decizia ta "
        "(MOQ, cubaj, lăţime război). Foloseşte celula B2.",
    ]


EXPORT_COLS = ["sku", "denumire_articol", "factory", "factory_source",
               "producer_entity", "family", "width_cm",
               "stock_units", "wh_units", "store_units", "unit_cost", "stock_value",
               "rate_mo", "months_of_supply", "days_of_cover", "units_90d", "units_365d",
               "safety_stock", "reorder_point", "target_stock", "suggested_qty",
               "suggested_value", "criticality", "colour", "substate", "product_state",
               "state_source", "lead_time_days", "lead_time_source", "months_since_sale",
               "simple_rotation_months", "cost_source"]


@app.get("/api/export")
def api_export(factory: str | None = None, family: str | None = None,
               colour: str | None = None, substate: str | None = None,
               reorder_only: bool = False, level: str = "custom"):
    p = P()
    bits = []
    if factory:
        p = p[p["factory"] == factory]
        bits.append(factory)
    if family:
        p = p[p["family"] == family]
        bits.append(family)
    if colour:
        p = p[p["colour"] == colour]
        bits.append(colour)
    if substate:
        p = p[p["substate"] == substate]
        bits.append(substate)
    if reorder_only:
        p = p[p["suggested_qty"] > 0]
        bits.append("de_comandat")
    if not len(p):
        raise HTTPException(404, "nothing to export")

    sort_col = "priority_score" if reorder_only else "stock_value"
    df = p.sort_values(sort_col, ascending=False)[EXPORT_COLS].copy()
    name = "_".join(["alarma", level] + [str(b).replace(" ", "-")[:24] for b in bits]) or "alarma"
    title = f"ALARMA V1 — {name} — fotografie {M()['as_of']}"
    data = _excel(df, sheet=name[:31], title=title,
                  qty_col="suggested_qty" if reorder_only else None,
                  notes=_export_notes())
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}.xlsx"'},
    )


# ------------------------------------------------------------------- email ---

@app.get("/api/trend")
def api_trend(months: int = 56):
    """Importer stock over time — the question nothing could answer before.

    The photograph is a single month, so "is the stock position improving?" was
    unanswerable in either direction. The code-keyed importer export carries 56
    months, which turns the north-star metric into something measurable instead of
    asserted.
    """
    f = paths.OUT_DIR / "stock_history.parquet"
    if not f.exists():
        return JSONResponse(_clean({"available": False,
                                    "reason": "no stock history — coded importer export missing"}))
    h = pd.read_parquet(f).sort_values("month").tail(months)
    rows = [{"month": str(pd.Timestamp(r["month"]).date()),
             "units": _f(r["units"]), "value": _f(r["value"]), "codes": int(r["codes"])}
            for _, r in h.iterrows()]
    first, last = rows[0], rows[-1]

    # Cohort stability. The series is NOT like-for-like: the range turns over, so a
    # value rise can be pure replacement rather than the same stock appreciating.
    # Publishing the deltas without this invites exactly the wrong conclusion.
    cohort = {"first_codes": first["codes"], "last_codes": last["codes"], "in_both": None}
    alt: dict = {}
    fsku = paths.OUT_DIR / "stock_history_sku.parquet"
    if fsku.exists():
        hs = pd.read_parquet(fsku)
        f_m, l_m = hs["month"].min(), hs["month"].max()
        a = set(hs.loc[hs["month"] == f_m, "sku"])
        b = set(hs.loc[hs["month"] == l_m, "sku"])
        cohort["in_both"] = len(a & b)
        by = hs.groupby("month").agg(u=("units", "sum"), v=("value", "sum"))
        for lbl, m0, m1 in (("aug22", "2022-08-01", "2026-08-01"),
                            ("aug25", "2025-08-01", "2026-08-01")):
            t0, t1 = pd.Timestamp(m0), pd.Timestamp(m1)
            if t0 in by.index and t1 in by.index:
                alt[f"{lbl}_units"] = _f((by.loc[t1, "u"] / by.loc[t0, "u"] - 1) * 100)
                alt[f"{lbl}_value"] = _f((by.loc[t1, "v"] / by.loc[t0, "v"] - 1) * 100)

    au0 = first["value"] / first["units"] if first["units"] else 0
    au1 = last["value"] / last["units"] if last["units"] else 0
    return JSONResponse(_clean({
        "available": True,
        "rows": rows,
        "first": first, "last": last,
        "delta_units_pct": _f((last["units"] / first["units"] - 1) * 100) if first["units"] else 0.0,
        "delta_value_pct": _f((last["value"] / first["value"] - 1) * 100) if first["value"] else 0.0,
        "delta_avg_cost_pct": _f((au1 / au0 - 1) * 100) if au0 else 0.0,
        "cohort": cohort,
        "alt_baselines": alt,
        "comparability_warning": (
            "Not a like-for-like cohort: the product range turns over between the "
            "endpoints, so a value change reflects replacement and mix, not the same "
            "stock appreciating. Do not read this as proof the position improved."),
    }))


@app.get("/api/cost-audit")
def api_cost_audit():
    """Every SKU whose stock value rests on an assumed cost, named.

    Three buckets: a real acquisition cost from the importer export, an assumed
    percentage of the realised selling price, or no value at all (contributes
    0 lei, so the capital total is understated). Downloadable as Excel so the
    assumed rows can be handed back for the real numbers.
    """
    p = P()
    cols = ["sku", "denumire_articol", "factory", "family", "stock_units",
            "wh_units", "store_units", "unit_cost", "avg_price", "stock_value",
            "cost_source", "colour"]
    df = p[cols].sort_values(["cost_source", "stock_value"], ascending=[True, False]).copy()
    notes = {
        "warehouse_export": "REAL cost — VALOARE STOC / STOC from the importer export",
        "assumed_pct_of_price": "ASSUMED — % of realised average selling price",
        "unknown": "NO VALUE — contributes 0 lei; capital total understated",
    }
    df["cost_basis_note"] = df["cost_source"].map(notes)
    data = _excel(df, sheet="cost_source_audit",
                  title=f"ALARMA V1 — audit sursa cost / cost source audit — {M()['as_of']}",
                  qty_col=None, notes=_export_notes())
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="alarma_audit_cost.xlsx"'},
    )


@app.get("/api/cost-summary")
def api_cost_summary():
    p = P()
    g = (p.groupby("cost_source")
         .agg(skus=("sku", "count"), units=("stock_units", "sum"),
              lei=("stock_value", "sum")).reset_index())
    tot_v = _f(p["stock_value"].sum())
    g["lei_share"] = np.where(tot_v > 0, g["lei"] / tot_v * 100, 0.0)
    return JSONResponse(_clean({"rows": g.to_dict("records"), "total_value": tot_v}))


_ORIGIN_RE = re.compile(r"^https?://[A-Za-z0-9.\-]+(:\d{1,5})?$")


def _safe_origin(value: str) -> str:
    """Only a bare http(s) origin is allowed through into the email HTML.

    `base_url` was interpolated straight into an href, so
    `?base_url=x"><img src=x onerror=alert(1)>` returned executable HTML on the
    app's own origin. Escaping alone would defuse this one payload; validating
    the shape is stronger and costs nothing, because the only real caller passes
    `location.origin`. Anything else falls back to the default.
    """
    v = (value or "").strip().rstrip("/")
    return v if _ORIGIN_RE.match(v) else ""


@app.get("/api/email", response_class=HTMLResponse)
def api_email(lang: str = "ro", base_url: str = ""):
    from app.email_render import render_email
    return HTMLResponse(render_email(P(), M(), _state["cfg"], lang,
                                     _safe_origin(base_url)))


# ------------------------------------------------------------------ static ---

@app.get("/")
def index():
    return FileResponse(WEB / "index.html")


app.mount("/static", StaticFiles(directory=WEB), name="static")


# --- DISPOSABLE: design experiments -----------------------------------------
# Serves alternative front-ends from design_experiments/<name>/web/ against this
# same API, so variants can be compared side by side with the original.
#
# Deliberately inert if the folder is absent: delete or archive
# alarm_v1/design_experiments/ and this block does nothing. Nothing else in the
# application reads from it, and no data, config or pipeline code lives there.
_EXPERIMENTS = Path(__file__).resolve().parents[1] / "design_experiments"

for _exp in sorted(_EXPERIMENTS.glob("*/web")) if _EXPERIMENTS.exists() else []:
    _name = _exp.parent.name

    def _exp_index(_dir=_exp):
        return FileResponse(_dir / "index.html")

    app.get(f"/design/{_name}", include_in_schema=False)(_exp_index)
    app.get(f"/design/{_name}/", include_in_schema=False)(_exp_index)
    app.mount(f"/design/{_name}/static", StaticFiles(directory=_exp), name=f"design_{_name}")


@app.on_event("startup")
def _startup():
    load_state()
